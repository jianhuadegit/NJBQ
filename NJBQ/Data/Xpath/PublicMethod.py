# author:zhujianhua
# contact: 996376627@qq.com
# datetime:2020\12\12
import random

from openpyxl import load_workbook,workbook,worksheet
from selenium import webdriver
from selenium.webdriver import ActionChains

from Data.Excel import rolefile
from Data.Xpath import Release_path
import time


#打开浏览器，获取网址，放大浏览器窗口
wd=webdriver.Chrome()
def OPenbrowser(weburl):
    wd.implicitly_wait(20)
    wd.get(weburl)
    wd.maximize_window()
#登录天府科技云
def loginTianfkjy(username,pwd):
    wd.find_element_by_css_selector(Release_path.logon_xpath).click()
    wd.find_element_by_css_selector(Release_path.username_xpath).send_keys(username)
    wd.find_element_by_css_selector(Release_path.pwd_xpath).send_keys(pwd)
    time.sleep(8)
    wd.find_element_by_css_selector(Release_path.loginbutton_xpath).click()
    print('欢迎登陆天府科技云',username)
def releaseReq():
#   点击买家中心
    wd.find_element_by_css_selector(Release_path.buy_xpath).click()
#   科技服务需求
    release = wd.find_element_by_css_selector(Release_path.Release_selector)
    ActionChains(wd).move_to_element(release).perform()
    Science_severce_class = wd.find_elements_by_class_name(Release_path.keji)
    Science_severce_class[0].click()
    hands = wd.window_handles
    hand = list(hands)
    wd.switch_to_window(hand[len(hand) - 1])
    wd.find_element_by_css_selector(Release_path.trade_selector).click()
    time.sleep(3)
    information = wd.find_element_by_css_selector(Release_path.information)
    ActionChains(wd).move_to_element(information).perform()
    ActionChains(wd).double_click(information).perform()
#   输入标题
    wd.find_element_by_css_selector(Release_path.logo).send_keys(getdatafromExcel()[0])
    #选择标签
    wd.find_element_by_css_selector(Release_path.select).click()
    time.sleep(1)
    lll=wd.find_element_by_css_selector(Release_path.development)
    time.sleep(1.5)
    lll.click()
#输入需求描述
    wd.find_element_by_css_selector(Release_path.describe).send_keys(addrole()+getdatafromExcel()[1])
#点击面议
    wd.find_element_by_css_selector(Release_path.checkbox).click()
#点击地区要求
    wd.find_element_by_css_selector(Release_path.region).click()
#点击全国
    all = wd.find_element_by_xpath(Release_path.allregion)
    ActionChains(wd).double_click(all).perform()
#点击我已阅读
    wd.find_element_by_css_selector(Release_path.read).click()
    #click releas
    # time.sleep(randomNum_sleeptime())
#点击发布
    wd.find_element_by_css_selector(Release_path.releabutton).click()
    # if check()!=None:
    #     i=0
    #     i+=i
    #     print(i+1,'条发布失败')
    # else:
    #     return
#检查必填
def check():
    text=wd.find_element_by_css_selector(Release_path.bitian).text
    return text
#产生发布之前的停顿随机数
def randomNum_sleeptime():
    sleeptime=random.randint(5,10)
    return sleeptime
#产生excel中的行标随机数
def randomNum_projectTitle():
    p = random.randint(1, 1916)
    return p
#从excel中拿数据出来
def getdatafromExcel():
    savepath = r'E:\NJBQ\Data\Excel\\' + r'testexcel1.xlsx'
    wb = load_workbook(savepath)
    ws=wb.active
    row = randomNum_projectTitle()
    title=ws.cell(row,1).value
    details=ws.cell(row,2).value
    wb.close()
    global titleanddetails
    titleanddetails=[title,details]
    print(row,titleanddetails[0])
    print(titleanddetails[1])
    return titleanddetails
#产生随机四句套话随机数
def randomNum_role():
    r=[random.randint(1,59)for i in range(4)]
    print(r)
    return r
#产生随机四句套话
def addrole():
    roledata = randomNum_role()
    addrole = []
    for i in range(4):
        addrole.append(rolefile.role[roledata[i]])
    roledata = addrole[0] + addrole[1] + addrole[2] + addrole[3]
    return roledata
#丢弃模块
def getregion():
#   点击名字
    wd.find_element_by_css_selector(Release_path.usernameimg).click()
#   用户中心
    wd.find_element_by_css_selector(Release_path.usercenter).click()
    # usercenter[0].click()
#   切换句柄
    hands = wd.window_handles
    hand = list(hands)
    wd.switch_to_window(hand[len(hand) - 1])
#   找到地区
    regiontext=wd.find_element_by_css_selector(Release_path.region1).text
    print(regiontext,'1')
#   返回地区





