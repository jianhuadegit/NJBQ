# author:zhujianhua
# contact: 996376627@qq.com
# datetime:2020/12/12
# software: PyCharm
import Case
from Data.Xpath import Huixinyun
from Data.Xpath.PublicMethod import wd
import  time
from openpyxl import Workbook,load_workbook,worksheet


#获取所需项目的信息
def GetData():
    wd.find_element_by_css_selector(Huixinyun.requsthome).click()
#抓取信息
def Getalldata(num,num1):
    firstData=wd.find_elements_by_class_name(Huixinyun.list0)
    if num==14:
        Gotonextpage()
    else:
        firstData[num].click()
        #获取项目名
        global firsttitle,firstdetails
        firsttitle=wd.find_element_by_css_selector(Huixinyun.list0title).text
        #获取需求详情
        firstdetails=wd.find_element_by_css_selector(Huixinyun.list0details).text
        print('第',num1,'页','第',num+1,'条','*******************',firsttitle)
        print(firstdetails)
    wd.back()
    time.sleep(1.5)
    #创建工作簿


    global wb
    global savepath3
    savepath3 = r'E:\NJBQ\Data\Excel\\' + r'testexcel2.xlsx'

    ws=wb.active

    ws.cell(14*(num1-1)+num+1,1,firsttitle)
    ws.cell(14*(num1-1)+num+1,2,firstdetails)
    save()
    print(14*(num1-1)+num+1,'已经OK')
def save():
    wb.save(savepath3)
def Openfile():
    savepath = r'E:\NJBQ\Data\Excel\\' + r'testexcel.xlsx'
    wb = load_workbook(savepath)
    return  wb
wb=Openfile()


def Gotonextpage():
    wd.find_element_by_partial_link_text('下一页').click()
    # time.sleep(1)
def Gotonextpage1():
    wd.find_element_by_partial_link_text('下一页').click()
    time.sleep(1)




